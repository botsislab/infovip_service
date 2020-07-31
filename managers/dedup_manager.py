from managers import qq_file_manager
from io import BytesIO
import db_util

# def get_total_pairs(number_of_items):
#     return choose(number_of_items, 2)

# def choose(n, k):
#     """
#     A fast way to calculate binomial coefficients by Andrew Dalke (contrib).
#     """
#     if 0 <= k <= n:
#         ntok = 1
#         ktok = 1
#         for t in range(1, min(k, n - k) + 1):
#             ntok *= n
#             ktok *= t
#             n -= 1
#         return ntok // ktok
#     else:
#         return 0

def get_deduped_file_bytes(case_list_id):
    from openpyxl import load_workbook
    from openpyxl.writer.excel import save_virtual_workbook
    from openpyxl.styles import Alignment
    from copy import copy

    # Load workbook
    case_list_file_bytes = qq_file_manager.get_case_list_file_bytes(case_list_id)
    workbook = load_workbook(case_list_file_bytes)

    case_details_sheet = workbook['Case Details']

    # Get offsets for new info
    reference_cell_contents = 'FAERS Case #'

    # Default to A1
    row_offset = 0
    column_offset = 0
    reference_cell = case_details_sheet['A1']
    if case_details_sheet['B2'].value == reference_cell_contents:
        row_offset = 1
        column_offset = 1
        reference_cell = case_details_sheet['B2']
    if case_details_sheet['B3'].value == reference_cell_contents:
        row_offset = 2
        column_offset = 1
        reference_cell = case_details_sheet['B3']
    
    dup_column_number = 3 + column_offset

    # Insert a new column on the case details sheet
    if case_details_sheet is None:
        raise(Exception('File in database did not contain a Case Details sheet'))

    case_details_sheet.insert_cols(dup_column_number)
    header_cell = case_details_sheet.cell(row = 1 + row_offset, column = dup_column_number, value = 'Suggested Duplicate Group')
    header_cell._style = copy(reference_cell._style)

    # Get suggested duplicate groups
    query = 'SELECT case_id, group_number FROM case_list_suggested_duplicate_groups WHERE case_list_id = %s'
    suggested_groups = db_util.get_all_as_dicts(query, (case_list_id,))
    dup_group_lookup = {suggested_group['case_id'] : suggested_group['group_number'] for suggested_group in suggested_groups}

    # Insert suggested group number for each row
    for row_index, row in enumerate(case_details_sheet.values):
        # Ignore rows that don't contain a case
        if row_index <= row_offset or row[column_offset] is None or row[column_offset + 1] is None:
            continue

        # Set cell's value
        case_id = str(row[column_offset]) + '-' + str(row[column_offset + 1])
        value = dup_group_lookup[case_id] if case_id in dup_group_lookup else None
        cell = case_details_sheet.cell(row = 1 + row_index, column = dup_column_number, value = value)

        # Set cell's style
        cell.border = copy(reference_cell.border)
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')

    deduped_file_bytes = BytesIO(save_virtual_workbook(workbook))
    
    return deduped_file_bytes